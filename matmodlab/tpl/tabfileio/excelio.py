#!/usr/bin/env python
import numpy as np

# For 'xls' reading
try:
    import xlrd
except ImportError:
    xlrd = None

# For 'xls' writing
try:
    import xlwt
except ImportError:
    xlwt = None

# For 'xlsx' reading/writing
try:
    import openpyxl
except ImportError:
    openpyxl = None


def read_excel(filename, sheetname=None, columns=None, disp=1):
    '''
    Takes in a spreadsheet of the form:

          A      B      C      D
      +------+------+------+------+
    1 | TIME | VAR1 | VAR2 | VAR3 |
      +------+------+------+------+
    2 |  0.0 |  0.0 |  0.0 |  0.0 |
      +------+------+------+------+
    3 |  1.0 |  0.5 |  0.0 |  0.0 |
      +------+------+------+------+
    4 |  2.0 |  1.0 |  0.5 |  0.0 |
      +------+------+------+------+
    5 |  3.0 |  0.5 |  1.0 |  0.0 |
      +------+------+------+------+
    6 |  4.0 |  0.0 |  0.5 |  0.0 |
      +------+------+------+------+
    7 |  5.0 |  0.0 |  0.0 |  0.0 |
      +------+------+------+------+

    And dumps it out into the following python format:

    head = ['TIME', 'VAR1', 'VAR2', 'VAR3']
    data = [[0.0, 0.0, 0.0, 0.0],
            [1.0, 0.5, 0.0, 0.0],
            [2.0, 1.0, 0.5, 0.0],
            [3.0, 0.5, 1.0, 0.0],
            [4.0, 0.0, 0.5, 0.0],
            [5.0, 0.0, 0.0, 0.0]]
    '''

    filetype = "XLS" if filename.upper().endswith(".XLS") else "XLSX"

    #
    # Open the file and grab the sheet names
    if filetype == "XLS":
        if xlrd is None:
            raise Exception("Cannot read xls files because package"
                            " xlrd is not found")
        wb = xlrd.open_workbook(filename)
        sheet_names = wb.sheet_names()
    else:
        if openpyxl is None:
            raise Exception("Cannot read xlsx files because package"
                            " openpyxl is not found")
        # Version checking
        major, minor, micro = [int(_) for _ in openpyxl.__version__.split(".")]
        if major < 2:
            raise Exception("Cannot read xlsx files because openpyxl package"
                            " version must be 2.0.0 or later. Found version"
                            " {0}".format(openpyxl.__version__))

        wb = openpyxl.load_workbook(filename, use_iterators=True,
                                    data_only=True)
        sheet_names = wb.get_sheet_names()

    #
    # Figure out what sheet to use
    if type(sheetname) is int:
        # Sheet by index
        if not (0 <= sheetname < len(sheet_names)):
            raise Exception('Sheet index "{0}" out of bounds. 0 <= idx <= {1}'
                            .format(sheetname, len(sheet_names)-1))
        sheet_index = sheetname
    elif type(sheetname) is str:
        # Sheet by name
        lower_names = [_.lower() for _ in sheet_names]
        if sheetname.lower() not in lower_names:
            raise Exception('Sheet "{0}" not found in {1}'
                            .format(sheetname.lower(), repr(sheet_names)))
        sheet_index = lower_names.index(sheetname.lower())
    else:
        # If sheet 'MML' exists, use it, otherwise grab the first one
        up = [_.upper() for _ in sheet_names]
        if 'MML' in up:
            sheet_index = up.index('MML')
        else:
            sheet_index = 0

    #
    # read in all the data
    if filetype == "XLS":
        sh = wb.sheet_by_index(sheet_index)
        head = list(map(str, sh.row_values(0)))
        head = sh.row_values(0)

        data = []
        row_index = 1
        while row_index < sh.nrows:
            # Make sure there is data for each cell under each header
            tmpdata = []
            for col_index in range(0, len(head)):
                # Cell type key:
                #   0 is blank; 1 is text; 2 is number; 3 is date
                if sh.cell_type(row_index, col_index) == 1:
                    tmpdata.append(float(sh.cell_value(row_index, col_index)))
                elif sh.cell_type(row_index, col_index) == 2:
                    tmpdata.append(sh.cell_value(row_index, col_index))
                else:
                    # irregular formatting. Exit, we're done!
                    break
            data.append(tmpdata)
            row_index += 1
    else:
        sh = wb.get_sheet_by_name(sheet_names[sheet_index])
        wholesheet = []
        for row in sh.iter_rows():
            tmprow = []
            for cell in row:
                tmprow.append(cell.value)
            wholesheet.append(tmprow)
        head = wholesheet[0]
        data = wholesheet[1:]

    data = np.array(data)

    # If specific columns are requested, filter the data
    if columns is not None:
        if any(isinstance(x, str) for x in columns):
            h = [s.lower() for s in head]
            for (i, item) in enumerate(columns):
                if isinstance(item, str):
                    columns[i] = h.index(item.lower())

        if head is not None:
            head = [head[i] for i in columns]
        data = data[:, columns]

    if not disp:
        return data
    return head, data


def write_excel(filename, head, data, columns=None, sheetname=None):

    filetype = "XLS" if filename.upper().endswith(".XLS") else "XLSX"
    sheetname = sheetname if sheetname else "Sheet1"

    if not isinstance(sheetname, str):
        raise Exception("kwarg 'sheetname' must be string-like.")

    #
    # Open the file and grab the sheet names
    if filetype == "XLS":
        if xlwt is None:
            raise Exception("Cannot write xls files because package"
                            " xlrd is not found")
        wb = xlwt.Workbook()
        sh = wb.add_sheet(sheetname)
    else:
        if openpyxl is None:
            raise Exception("Cannot write xlsx files because package"
                            " openpyxl is not found")
        wb = openpyxl.Workbook()
        sh = wb.get_active_sheet()
        sh.title = sheetname

    # If specific columns are requested, prepare the columns index
    if columns is not None:
        if any(isinstance(x, str) for x in columns):
            h = [s.lower() for s in head]
            for (i, item) in enumerate(columns):
                if isinstance(item, str):
                    columns[i] = h.index(item.lower())
    else:
        columns = list(range(0, len(head)))

    # Write to the worksheet
    if filetype == "XLS":
        # write the headers
        for idx, i in enumerate(columns):
            sh.write(0, idx, head[i])
        # write the data
        for rdx, row in enumerate(data):
            for idx, i in enumerate(columns):
                sh.write(rdx + 1, idx, row[i])
        wb.save(filename)
    else:
        # write the headers
        sh.append([head[i] for i in columns])
        # write the data
        for row in data:
            sh.append([row[i] for i in columns])
        wb.save(filename=filename)


if __name__ == '__main__':
    cols = ["TIME", "STRESS", 1]

    print("===== loading xls")
    xlshead, xlsdata = read_excel("io_test.xls", columns=cols)
    print("===== loading xlsx")
    xlsxhead, xlsxdata = read_excel("io_test.xlsx", columns=cols)

    print("===== HEADERS")
    print("xls headers: " + repr(xlshead))
    print("xlsx headers:" + repr(xlsxhead))
    print("are headers the same?" + repr(xlshead == xlsxhead))
    print("===== DATA")
    print("xls data:  " + repr(xlsdata))
    print("xlsx data: " + repr(xlsxdata))
    print("data diff: " + repr(xlsdata - xlsxdata))
